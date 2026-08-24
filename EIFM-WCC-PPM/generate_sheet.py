
from pathlib import Path
from copy import copy
from datetime import datetime
from io import BytesIO

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.page import PageMargins

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics

BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_DIR = BASE_DIR / "templates"
ASSET_DIR = BASE_DIR / "assets"

# Exact equipment sheet names from the two supplied Excel templates.
def _sheet_names(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    return list(wb.sheetnames)

EQUIPMENT_SHEETS = list(dict.fromkeys(
    _sheet_names(TEMPLATE_DIR / "All-3.xlsx") +
    _sheet_names(TEMPLATE_DIR / "All.xlsx")
))


def _copy_sheet_from_workbook(source_path, sheet_name):
    # Load the supplied workbook itself, remove every unrelated sheet,
    # and retain the selected equipment sheet. This preserves the original
    # workbook's formatting, merged cells, dimensions, print setup, formulas,
    # row heights and column widths as far as openpyxl supports.
    wb = openpyxl.load_workbook(source_path)
    keep = wb[sheet_name]
    for name in list(wb.sheetnames):
        if name != sheet_name:
            del wb[name]
    return wb, keep


def _set_if_merged(ws, cell_ref, value):
    for merged in ws.merged_cells.ranges:
        if cell_ref in merged:
            ws.cell(merged.min_row, merged.min_col).value = value
            return
    ws[cell_ref] = value


def _fill_equipment_metadata(ws, meta):
    # The supplied sheets use C3:I3, C4:I4 ... for the metadata area.
    mapping = {
        "C3": meta["project"],
        "C4": meta["location"],
        "C5": meta["unit_number"],
        "C6": meta["frequency"],
        "C7": meta["category"],
        "C8": meta["equipment"],
        "K3": meta["fiscal_year"],
        "K4": meta["wo_number"],
        "K5": meta["scheduled_month"],
        "K6": meta["service_date"].strftime("%d/%m/%Y"),
        "K7": meta["time_start"].strftime("%H:%M"),
        "K8": meta["time_finish"].strftime("%H:%M"),
    }
    for ref, value in mapping.items():
        _set_if_merged(ws, ref, value)


def _style_ppm_front(ws):
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.35, right=0.35, top=0.35, bottom=0.35)

    for col, width in {"A": 5, "B": 21, "C": 4, "D": 25, "E": 4, "F": 18, "G": 18, "H": 18}.items():
        ws.column_dimensions[col].width = width

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:H1")
    ws["A1"] = "EIFM"
    ws["A1"].font = Font(name="Arial", size=20, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = "PLANNED PREVENTIVE MAINTENANCE"
    ws["A2"].font = Font(name="Arial", size=15, bold=True, underline="single")
    ws["A2"].alignment = Alignment(horizontal="center")

    rows = [
        ("1", "PPM Number", "ppm_number"),
        ("2", "Client", "client"),
        ("3", "Project", "project"),
        ("4", "Location", "location"),
        ("5", "Details of Work", "details_of_work"),
        ("6", "Date & Time of Service", "service_datetime"),
        ("7", "Unit Number", "unit_number"),
        ("8", "Equipment", "equipment"),
        ("9", "Frequency", "frequency"),
        ("10", "Category", "category"),
        ("11", "WO Number", "wo_number"),
    ]

    r = 4
    for num, label, key in rows:
        ws[f"A{r}"] = num + "."
        ws[f"A{r}"].font = Font(bold=True)
        ws[f"B{r}"] = label
        ws[f"B{r}"].font = Font(bold=True)
        ws[f"C{r}"] = ":-"
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
        ws[f"D{r}"].border = border
        ws[f"D{r}"].alignment = Alignment(vertical="center", wrap_text=True)
        r += 1

    # Work description gets more room, matching the WCC concept.
    ws.row_dimensions[8].height = 55
    ws["D8"].alignment = Alignment(vertical="top", wrap_text=True)

    # Equipment / schedule summary.
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws[f"A{r}"] = "PPM CHECK LIST / EQUIPMENT"
    ws[f"A{r}"].font = Font(bold=True, size=12)
    ws[f"A{r}"].alignment = Alignment(horizontal="center")
    ws[f"A{r}"].fill = PatternFill("solid", fgColor="D9EAD3")
    return rows


def generate_ppm_workbook(source_path, equipment, meta, output_path):
    wb, equipment_ws = _copy_sheet_from_workbook(source_path, equipment)

    # Create the front page in the same workbook; the selected equipment sheet
    # remains the original supplied Excel sheet.
    front = wb.create_sheet("PPM Front Page", 0)
    rows = _style_ppm_front(front)

    service_datetime = (
        meta["service_date"].strftime("%d/%m/%Y")
        + " "
        + meta["time_start"].strftime("%H:%M")
        + " - "
        + meta["time_finish"].strftime("%H:%M")
    )
    values = {
        "ppm_number": meta["ppm_number"],
        "client": meta["client"],
        "project": meta["project"],
        "location": meta["location"],
        "details_of_work": meta["details_of_work"],
        "service_datetime": service_datetime,
        "unit_number": meta["unit_number"],
        "equipment": meta["equipment"],
        "frequency": meta["frequency"],
        "category": meta["category"],
        "wo_number": meta["wo_number"],
    }
    start = 4
    for _, _, key in rows:
        front[f"D{start}"] = values[key]
        start += 1

    # Put the supplied EIFM logo on the front page.
    logo_path = ASSET_DIR / "EIFM_logo.jpg"
    if logo_path.exists():
        img = XLImage(str(logo_path))
        img.width = 65
        img.height = 108
        front.add_image(img, "F1")

    # Fill only the existing metadata cells in the supplied checklist.
    _fill_equipment_metadata(equipment_ws, meta)

    # Make the selected checklist sheet active after generation.
    wb.active = 0
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _font():
    # DejaVu has broad Unicode coverage for the PDF text.
    candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans.ttf",
    ]
    for p in candidates:
        if Path(p).exists():
            pdfmetrics.registerFont(TTFont("EIFM", p))
            return "EIFM"
    return "Helvetica"


def _line(c, x1, y, x2):
    c.line(x1, y, x2, y)


def _field(c, y, label, value, x_label=70, x_value=210, width=390):
    c.setFont(FONT, 10.5)
    c.drawString(x_label, y, label)
    c.drawString(x_value, y, ":-")
    _line(c, x_value + 15, y - 2, x_value + width)
    if value:
        c.drawString(x_value + 20, y, str(value)[:55])


def generate_wcc_pdf(meta, output_path):
    c = canvas.Canvas(str(output_path), pagesize=A4)
    W, H = A4
    c.setTitle("EIFMEN08 Work Completion Certificate")
    c.setFont(FONT, 10)

    # Outer border matching the supplied WCC.
    c.rect(30, 28, W - 60, H - 56)

    # Logo
    logo = ASSET_DIR / "EIFM_logo.jpg"
    if logo.exists():
        c.drawImage(str(logo), W/2 - 35, H - 115, width=70, height=116, preserveAspectRatio=True, mask="auto")

    c.setFont(FONT, 16)
    c.setFont(FONT, 16)
    c.drawCentredString(W/2, H - 135, "WORK COMPLETION CERTIFICATE")

    y = H - 165
    _field(c, y, "1.   Job Order Number", meta.get("job_order", ""))
    y -= 25
    _field(c, y, "2.   Client", meta.get("client", ""))
    y -= 25
    _field(c, y, "3.   Project", meta.get("project", ""))
    y -= 25
    _field(c, y, "4.   Location", meta.get("location", ""))
    if meta.get("tel_no"):
        c.drawString(445, y, "Tel. No. :-")
        c.drawString(505, y, str(meta["tel_no"])[:18])

    y -= 28
    c.setFont(FONT, 10.5)
    c.drawString(70, y, "5.   Details of Work :-")
    y -= 22

    details = meta.get("details", "") or ""
    max_chars = 92
    words = details.split()
    lines = []
    line = ""
    for word in words:
        if len(line) + len(word) + 1 <= max_chars:
            line = (line + " " + word).strip()
        else:
            lines.append(line)
            line = word
    if line:
        lines.append(line)
    lines = lines[:6]
    for _ in range(6):
        if lines:
            txt = lines.pop(0)
            if txt:
                c.drawString(85, y, txt)
        c.setDash(1, 2)
        _line(c, 85, y - 8, W - 85)
        c.setDash()
        y -= 26

    y -= 5
    completion = f"{meta.get('completion_date','')} {meta.get('completion_time','')}"
    _field(c, y, "6.   Date & Time of Completion", completion, x_value=300, width=285)

    y -= 33
    c.drawString(70, y, "7.   Signature of Site in Charge:-")
    _line(c, 300, y - 2, 405)
    c.drawString(420, y, "Name :-")
    _line(c, 468, y - 2, 540)
    c.drawString(550, y, "Date :-")
    _line(c, 590, y - 2, 580 + 50)
    y -= 23
    c.drawString(420, y, "ID :-")
    _line(c, 445, y - 2, 540)

    y -= 28
    c.drawString(70, y, "8.   Signature of HOD :-")
    _line(c, 205, y - 2, 350)
    c.drawString(365, y, "Name:-")
    _line(c, 410, y - 2, 500)
    c.drawString(510, y, "Date :-")
    _line(c, 555, y - 2, 590)
    y -= 23
    c.drawString(365, y, "ID :-")
    _line(c, 390, y - 2, 500)

    y -= 35
    c.drawString(70, y, "9.   Enclosed following documents with the work completion form:")
    y -= 22
    docs = ["LPO", "Invoice", "Delivery Note", "Petty Cash", "Material Requisition", "Job Completion"]
    x_positions = [105, 220, 350, 485, 145, 345]
    y_positions = [y, y, y, y, y - 28, y - 28]
    for idx, doc in enumerate(docs):
        x = x_positions[idx]
        yy = y_positions[idx]
        c.drawString(x, yy, f"{idx+1}.  {doc}")
        c.rect(x + (42 if idx < 4 else 82), yy - 3, 12, 12)

    y -= 75
    c.drawString(70, y, "Client Signature:-")
    _line(c, 165, y - 2, 320)
    y -= 28
    c.drawString(70, y, "Name :-")
    _line(c, 115, y - 2, 320)
    y -= 28
    c.drawString(70, y, "Phone No. :-")
    _line(c, 145, y - 2, 320)

    y -= 35
    c.setFont(FONT, 10)
    c.drawString(70, y, "Dear valued customer, please mark your satisfaction level in the scale below:")
    y -= 20
    c.rect(65, y - 12, W - 130, 32)
    levels = ["1. Poor", "2. Satisfied", "3. Good", "4. Very Good", "5. Excellent"]
    x = 95
    for level in levels:
        c.drawString(x, y, level)
        x += 93

    y -= 42
    c.setFont(FONT, 10.5)
    c.drawString(70, y, "Remarks /Suggestions:")
    c.rect(60, y - 125, W - 120, 105)

    c.drawString(75, y - 105, "Client Signature  --------------------")
    c.drawString(340, y - 105, "Date :- ____/____/_____")

    c.setFont(FONT, 8.5)
    c.drawCentredString(W/2, 47, "P.O Box 2286, Abu Dhabi - United Arab Emirates - Tel: +971-2-6436663, Fax: +971-2-6436660")
    c.drawCentredString(W/2, 34, "E-mail: eifm@eifm.ae     Website: www.eifm.ae")
    c.drawString(42, 17, "Page 1 of 1")
    c.drawCentredString(W/2, 17, "EIFMEN08")
    c.drawRightString(W - 42, 17, "Version R3")
    c.save()


FONT = _font()
