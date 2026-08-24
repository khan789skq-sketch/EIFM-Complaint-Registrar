
import os
import sqlite3
import hashlib
import secrets
from pathlib import Path
from datetime import datetime, date, time
from io import BytesIO

import streamlit as st
import openpyxl

from generate_sheet import (
    EQUIPMENT_SHEETS,
    generate_ppm_workbook,
    generate_wcc_pdf,
)

BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_DIR = BASE_DIR / "templates"
ASSET_DIR = BASE_DIR / "assets"
OUTPUT_DIR = BASE_DIR / "outputs"
DATA_DIR = BASE_DIR / "data"
DB_PATH = DATA_DIR / "eifm_app.db"

OUTPUT_DIR.mkdir(exist_ok=True)
DATA_DIR.mkdir(exist_ok=True)

st.set_page_config(
    page_title="EIFM PPM & WCC Generator",
    page_icon="EIFM",
    layout="wide",
)

st.markdown("""
<style>
.stApp {
    background: linear-gradient(135deg, #f7faf8 0%, #eef5f0 50%, #ffffff 100%);
}
section[data-testid="stSidebar"] {
    background: #0f5132;
}
section[data-testid="stSidebar"] * {
    color: white !important;
}
div[data-testid="stMetric"] {
    background: white;
    border: 1px solid #d8e4dc;
    border-radius: 12px;
    padding: 10px;
}
</style>
""", unsafe_allow_html=True)


def db():
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    return con


def init_db():
    with db() as con:
        con.execute("""
            CREATE TABLE IF NOT EXISTS users (
                email TEXT PRIMARY KEY,
                password_hash TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
        """)
        con.execute("""
            CREATE TABLE IF NOT EXISTS records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                email TEXT NOT NULL,
                kind TEXT NOT NULL,
                project TEXT,
                equipment TEXT,
                number TEXT,
                created_at TEXT NOT NULL,
                ppm_file TEXT,
                wcc_file TEXT
            )
        """)
        con.commit()


def hash_password(password: str) -> str:
    salt = secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac(
        "sha256", password.encode("utf-8"), salt, 200_000
    )
    return salt.hex() + "$" + digest.hex()


def verify_password(password: str, stored: str) -> bool:
    try:
        salt_hex, digest_hex = stored.split("$", 1)
        salt = bytes.fromhex(salt_hex)
        digest = hashlib.pbkdf2_hmac(
            "sha256", password.encode("utf-8"), salt, 200_000
        )
        return secrets.compare_digest(digest.hex(), digest_hex)
    except Exception:
        return False


def authenticate(email: str, password: str) -> bool:
    with db() as con:
        row = con.execute(
            "SELECT password_hash FROM users WHERE email = ?", (email.lower().strip(),)
        ).fetchone()
    return bool(row and verify_password(password, row["password_hash"]))


def create_user(email: str, password: str):
    email = email.lower().strip()
    if "@" not in email or "." not in email.split("@")[-1]:
        return False, "Enter a valid email address."
    if len(password) < 6:
        return False, "Password must be at least 6 characters."
    try:
        with db() as con:
            con.execute(
                "INSERT INTO users(email, password_hash, created_at) VALUES(?,?,?)",
                (email, hash_password(password), datetime.now().isoformat(timespec="seconds")),
            )
            con.commit()
        return True, "Account created. You can sign in now."
    except sqlite3.IntegrityError:
        return False, "This email is already registered."


def add_record(email, kind, project, equipment, number, ppm_file=None, wcc_file=None):
    with db() as con:
        con.execute("""
            INSERT INTO records
            (email, kind, project, equipment, number, created_at, ppm_file, wcc_file)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            email, kind, project, equipment, number,
            datetime.now().isoformat(timespec="seconds"),
            ppm_file, wcc_file
        ))
        con.commit()


def get_records(email):
    with db() as con:
        return con.execute(
            "SELECT * FROM records WHERE email=? ORDER BY id DESC", (email,)
        ).fetchall()


def equipment_source(equipment):
    all3 = TEMPLATE_DIR / "All-3.xlsx"
    all1 = TEMPLATE_DIR / "All.xlsx"
    if all3.exists():
        wb = openpyxl.load_workbook(all3, read_only=True)
        if equipment in wb.sheetnames:
            return all3
    if all1.exists():
        wb = openpyxl.load_workbook(all1, read_only=True)
        if equipment in wb.sheetnames:
            return all1
    raise FileNotFoundError(f"Equipment template not found: {equipment}")


init_db()

if "user" not in st.session_state:
    st.session_state.user = None

if not st.session_state.user:
    st.title("EIFM PPM & WCC Generator")
    st.caption("Planned Preventive Maintenance / Work Completion Certificate")
    tab1, tab2 = st.tabs(["Sign in", "Sign up"])

    with tab1:
        email = st.text_input("Email", key="login_email")
        password = st.text_input("Password", type="password", key="login_password")
        if st.button("Sign in", type="primary"):
            if authenticate(email, password):
                st.session_state.user = email.lower().strip()
                st.rerun()
            else:
                st.error("Invalid email or password.")

    with tab2:
        email2 = st.text_input("Email", key="signup_email")
        p1 = st.text_input("Password", type="password", key="signup_password")
        p2 = st.text_input("Confirm password", type="password", key="signup_confirm")
        if st.button("Create account"):
            if p1 != p2:
                st.error("Passwords do not match.")
            else:
                ok, msg = create_user(email2, p1)
                (st.success if ok else st.error)(msg)
    st.stop()

user = st.session_state.user

with st.sidebar:
    st.image(str(ASSET_DIR / "EIFM_logo.jpg"), width=105)
    st.markdown("### EIFM PPM & WCC")
    menu = st.radio("Menu", ["Dashboard", "New PPM", "New WCC", "My Records"])
    st.divider()
    st.caption(user)
    if st.button("Sign out"):
        st.session_state.user = None
        st.rerun()

if menu == "Dashboard":
    records = get_records(user)
    st.title("EIFM PPM & WCC Generator")
    c1, c2, c3 = st.columns(3)
    c1.metric("My Records", len(records))
    c2.metric("Equipment Templates", len(EQUIPMENT_SHEETS))
    c3.metric("PPM Excel Templates", 2)

    st.markdown("### Current PPM structure")
    st.info(
        "PPM numbering is 1st PPM through 4th PPM. "
        "The front page remains the same structure as the WCC front page, "
        "with the Job Order field changed to PPM Number. "
        "The full work description is: "
        "**Planned Preventive Maintenance Service Complete as per Attached Check List**. "
        "The selected equipment checklist is taken directly from the supplied Excel template."
    )

elif menu == "New PPM":
    st.title("New PPM")
    st.caption(
        "1st PPM / 2nd PPM / 3rd PPM / 4th PPM — same PPM front-page structure, "
        "with the selected Excel equipment checklist inside the generated workbook."
    )

    with st.form("ppm_form"):
        c1, c2 = st.columns(2)
        with c1:
            project = st.text_input("Project Name")
            location = st.text_input("Location")
            unit_number = st.text_input("Unit Number")
            category = st.text_input("Category")
            equipment = st.selectbox("Equipment", EQUIPMENT_SHEETS)
            frequency = st.selectbox(
                "Frequency",
                ["Monthly", "Quarterly", "Semi-Annual", "Annual", "Corrective / Complaint"]
            )
        with c2:
            client = st.text_input("Client")
            tel_no = st.text_input("Tel. No.")
            ppm_number = st.selectbox("PPM Number", ["1st PPM", "2nd PPM", "3rd PPM", "4th PPM"])
            fiscal_year = st.number_input("Fiscal Year", min_value=2000, max_value=2100, value=date.today().year)
            wo_number = st.text_input("WO Number")
            scheduled_month = st.selectbox(
                "Scheduled Month",
                ["January","February","March","April","May","June",
                 "July","August","September","October","November","December"]
            )
            service_date = st.date_input("Date of Service", value=date.today())
            time_start = st.time_input("Time Start", value=time(8, 0))
            time_finish = st.time_input("Time Finish", value=time(17, 0))

        st.markdown("**Details of Work**")
        st.text_area(
            "PPM Work Description",
            value="Planned Preventive Maintenance Service Complete as per Attached Check List",
            disabled=True,
            label_visibility="collapsed",
        )

        submitted = st.form_submit_button("Generate PPM", type="primary")

    if submitted:
        if not project.strip() or not location.strip() or not equipment:
            st.error("Project, Location and Equipment are required.")
        else:
            source = equipment_source(equipment)
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_project = "".join(ch if ch.isalnum() or ch in "._-" else "_" for ch in project)[:50]
            filename = f"PPM_{ppm_number.replace(' ', '_')}_{safe_project}_{stamp}.xlsx"
            out = OUTPUT_DIR / filename

            meta = {
                "client": client,
                "project": project,
                "location": location,
                "tel_no": tel_no,
                "unit_number": unit_number,
                "category": category,
                "equipment": equipment,
                "frequency": frequency,
                "ppm_number": ppm_number,
                "fiscal_year": fiscal_year,
                "wo_number": wo_number,
                "scheduled_month": scheduled_month,
                "service_date": service_date,
                "time_start": time_start,
                "time_finish": time_finish,
                "details_of_work": "Planned Preventive Maintenance Service Complete as per Attached Check List",
            }

            generate_ppm_workbook(source, equipment, meta, out)
            add_record(user, "PPM", project, equipment, ppm_number, ppm_file=str(out))
            st.success("PPM generated successfully.")
            st.download_button(
                "Download PPM Excel",
                data=out.read_bytes(),
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

elif menu == "New WCC":
    st.title("New WCC")
    st.caption("EIFMEN08 Work Completion Certificate — Version R3 structure.")

    with st.form("wcc_form"):
        c1, c2 = st.columns(2)
        with c1:
            job_order = st.text_input("Job Order Number")
            client = st.text_input("Client")
            project = st.text_input("Project")
            location = st.text_input("Location")
            tel_no = st.text_input("Tel. No.")
            details = st.text_area("Details of Work")
            completion_date = st.date_input("Date of Completion", value=date.today())
            completion_time = st.time_input("Time of Completion", value=time(17, 0))
        with c2:
            site_name = st.text_input("Site in Charge Name")
            site_id = st.text_input("Site in Charge ID")
            site_date = st.date_input("Site in Charge Date", value=date.today())
            hod_name = st.text_input("HOD Name")
            hod_id = st.text_input("HOD ID")
            hod_date = st.date_input("HOD Date", value=date.today())
            client_name = st.text_input("Client Signature Name")
            phone_no = st.text_input("Client Phone No.")
            remarks = st.text_area("Remarks / Suggestions")

        st.markdown("### Enclosed Documents")
        docs = st.multiselect(
            "Select enclosed documents",
            ["LPO", "Invoice", "Delivery Note", "Petty Cash", "Material Requisition", "Job Completion"],
        )
        satisfaction = st.selectbox(
            "Customer Satisfaction",
            ["Not marked", "1. Poor", "2. Satisfied", "3. Good", "4. Very Good", "5. Excellent"],
        )

        submitted = st.form_submit_button("Generate WCC", type="primary")

    if submitted:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_project = "".join(ch if ch.isalnum() or ch in "._-" else "_" for ch in project)[:50]
        filename = f"WCC_{job_order or 'NO_JOB'}_{safe_project}_{stamp}.pdf"
        out = OUTPUT_DIR / filename

        meta = {
            "job_order": job_order,
            "client": client,
            "project": project,
            "location": location,
            "tel_no": tel_no,
            "details": details,
            "completion_date": completion_date,
            "completion_time": completion_time,
            "site_name": site_name,
            "site_id": site_id,
            "site_date": site_date,
            "hod_name": hod_name,
            "hod_id": hod_id,
            "hod_date": hod_date,
            "client_name": client_name,
            "phone_no": phone_no,
            "remarks": remarks,
            "docs": docs,
            "satisfaction": satisfaction,
        }

        generate_wcc_pdf(meta, out)
        add_record(user, "WCC", project, "", job_order, wcc_file=str(out))
        st.success("WCC generated successfully.")
        st.download_button(
            "Download WCC PDF",
            data=out.read_bytes(),
            file_name=filename,
            mime="application/pdf",
        )

else:
    st.title("My Records")
    rows = get_records(user)
    if not rows:
        st.info("No records generated yet.")
    else:
        for row in rows:
            with st.container(border=True):
                st.write(
                    f"**{row['kind']}**  |  {row['project'] or '-'}  |  "
                    f"{row['equipment'] or '-'}  |  {row['number'] or '-'}"
                )
                st.caption(row["created_at"])
                if row["ppm_file"] and Path(row["ppm_file"]).exists():
                    p = Path(row["ppm_file"])
                    st.download_button(
                        "Download PPM",
                        data=p.read_bytes(),
                        file_name=p.name,
                        key=f"ppm_{row['id']}",
                    )
                if row["wcc_file"] and Path(row["wcc_file"]).exists():
                    p = Path(row["wcc_file"])
                    st.download_button(
                        "Download WCC",
                        data=p.read_bytes(),
                        file_name=p.name,
                        key=f"wcc_{row['id']}",
                    )
